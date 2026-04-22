"""Spreadsheet extraction to structured markdown.

Converts .xlsx, .xls, .csv, and .tsv files into markdown tables suitable
for LLM evidence extraction. Spreadsheets are already structured data, so
unlike PDFs there is no quality tradeoff — one extraction pass suffices.

Output format uses sheet markers that mirror the PDF page marker convention
(``--- Sheet "name" (N of M) ---``) so citation extraction works uniformly.
"""

import csv
import io
import logging
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from ..models.errors import DocumentExtractionError

logger = logging.getLogger(__name__)

MAX_ROWS_PER_SHEET = 10_000

# Char caps that keep LLM payloads under TELUS Cloudflare gateway's edge timeout.
# Observed behaviour: prompts approaching 100K chars (≈25K tokens) routinely
# trip 504/524 on the GPT-OSS-120B endpoint because inference exceeds the edge
# idle budget. Capping well below that ceiling keeps per-document prompts fast
# enough to complete while preserving representative evidence rows.
#
# Two caps apply. The per-sheet cap protects against one huge sheet blowing the
# budget alone; the per-workbook cap protects against the multi-sheet case
# where individually compliant sheets still sum to an oversized total.
MAX_CHARS_PER_SHEET = 40_000
MAX_CHARS_PER_WORKBOOK = 60_000


def _sanitize_cell(value: Any) -> str:
    """Convert a cell value to a pipe-safe markdown string."""
    if value is None:
        return ""
    text = str(value).strip()
    # Escape pipe characters so they don't break the markdown table
    text = text.replace("|", "\\|")
    # Collapse internal newlines to spaces
    text = re.sub(r"\s*\n\s*", " ", text)
    return text


def _rows_to_markdown(headers: list[str], rows: list[list[str]]) -> str:
    """Render a list of header + data rows as a markdown table."""
    if not headers:
        return ""

    safe_headers = [_sanitize_cell(h) or f"col_{i}" for i, h in enumerate(headers)]
    lines = [
        "| " + " | ".join(safe_headers) + " |",
        "| " + " | ".join("---" for _ in safe_headers) + " |",
    ]
    for row in rows:
        # Pad or truncate row to match header count
        padded = row + [""] * (len(safe_headers) - len(row))
        safe = [_sanitize_cell(c) for c in padded[: len(safe_headers)]]
        lines.append("| " + " | ".join(safe) + " |")

    return "\n".join(lines)


def _truncate_table_by_chars(
    headers: list[str],
    rows: list[list[str]],
    cap: int,
) -> tuple[str, int]:
    """Binary-search the largest row count whose rendered table fits under ``cap``.

    Returns (markdown, rows_kept). Called only when the full render exceeds
    ``cap``; the footer caller adds the truncation annotation.
    """
    lo, hi = 0, len(rows)
    best_md = _rows_to_markdown(headers, [])
    best_kept = 0
    while lo <= hi:
        mid = (lo + hi) // 2
        md = _rows_to_markdown(headers, rows[:mid])
        if len(md) <= cap:
            best_md = md
            best_kept = mid
            lo = mid + 1
        else:
            hi = mid - 1
    return best_md, best_kept


def _extract_xlsx(filepath: Path) -> tuple[str, int, int, int]:
    """Extract all sheets from an .xlsx/.xls workbook.

    Returns:
        (markdown, sheet_count, total_rows, tables_found)
    """
    try:
        import openpyxl
    except ImportError:
        raise DocumentExtractionError(
            "openpyxl not installed. Install with: uv pip install openpyxl",
            details={"package": "openpyxl"},
        )

    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    except Exception as e:
        raise DocumentExtractionError(
            f"Failed to open workbook: {e}",
            details={"filepath": str(filepath), "error": str(e)},
        )

    sheet_names = wb.sheetnames
    sheet_count = len(sheet_names)
    sections: list[str] = []
    total_rows = 0
    tables_found = 0

    for idx, name in enumerate(sheet_names, 1):
        ws = wb[name]

        # Read all rows (generator → list)
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append(list(row))

        if not all_rows:
            sections.append(f'--- Sheet "{name}" ({idx} of {sheet_count}) ---\n\n*Empty sheet*')
            continue

        # First non-empty row is treated as the header
        headers = all_rows[0]
        data_rows = all_rows[1:]

        truncated = False
        if len(data_rows) > MAX_ROWS_PER_SHEET:
            data_rows = data_rows[:MAX_ROWS_PER_SHEET]
            truncated = True

        total_rows += len(data_rows)
        tables_found += 1

        table_md = _rows_to_markdown(headers, data_rows)
        char_truncated_count: int | None = None
        if len(table_md) > MAX_CHARS_PER_SHEET:
            table_md, char_truncated_count = _truncate_table_by_chars(
                headers, data_rows, MAX_CHARS_PER_SHEET
            )
        section = f'--- Sheet "{name}" ({idx} of {sheet_count}) ---\n\n{table_md}'
        if truncated:
            section += (
                f"\n\n*Truncated by row cap: showing first {MAX_ROWS_PER_SHEET:,} "
                f"of {len(all_rows) - 1:,} data rows*"
            )
        if char_truncated_count is not None:
            section += (
                f"\n\n*Truncated by char cap ({MAX_CHARS_PER_SHEET:,}): kept "
                f"{char_truncated_count:,} of {len(data_rows):,} rows on this sheet*"
            )
        sections.append(section)

    wb.close()
    markdown = "\n\n".join(sections)
    return markdown, sheet_count, total_rows, tables_found


def _extract_csv(filepath: Path) -> tuple[str, int, int, int]:
    """Extract a .csv or .tsv file.

    Returns:
        (markdown, sheet_count=1, total_rows, tables_found)
    """
    delimiter = "\t" if filepath.suffix.lower() == ".tsv" else ","
    name = filepath.name

    try:
        text = filepath.read_text(encoding="utf-8", errors="replace")
    except Exception as e:
        raise DocumentExtractionError(
            f"Failed to read CSV file: {e}",
            details={"filepath": str(filepath), "error": str(e)},
        )

    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    all_rows = list(reader)

    if not all_rows:
        markdown = f'--- Sheet "{name}" (1 of 1) ---\n\n*Empty file*'
        return markdown, 1, 0, 0

    headers = all_rows[0]
    data_rows = all_rows[1:]

    truncated = False
    if len(data_rows) > MAX_ROWS_PER_SHEET:
        data_rows = data_rows[:MAX_ROWS_PER_SHEET]
        truncated = True

    table_md = _rows_to_markdown(headers, data_rows)
    char_truncated_count: int | None = None
    if len(table_md) > MAX_CHARS_PER_SHEET:
        table_md, char_truncated_count = _truncate_table_by_chars(
            headers, data_rows, MAX_CHARS_PER_SHEET
        )
    section = f'--- Sheet "{name}" (1 of 1) ---\n\n{table_md}'
    if truncated:
        section += (
            f"\n\n*Truncated by row cap: showing first {MAX_ROWS_PER_SHEET:,} "
            f"of {len(all_rows) - 1:,} data rows*"
        )
    if char_truncated_count is not None:
        section += (
            f"\n\n*Truncated by char cap ({MAX_CHARS_PER_SHEET:,}): kept "
            f"{char_truncated_count:,} of {len(data_rows):,} rows on this file*"
        )

    return section, 1, len(data_rows), 1 if data_rows else 0


async def extract_spreadsheet(filepath: str) -> dict[str, Any]:
    """Convert a spreadsheet file to structured markdown.

    Supports .xlsx, .xls, .csv, and .tsv. Each sheet becomes a markdown table
    under a sheet marker that mirrors the PDF page marker convention.

    Args:
        filepath: Path to spreadsheet file

    Returns:
        Dictionary with:
            - filepath: str — original file path
            - markdown: str — full markdown with sheet markers
            - sheet_count: int — number of sheets
            - total_chars: int — character count of markdown
            - tables_found: int — number of non-empty tables
            - extracted_at: str — ISO timestamp
            - extraction_method: str — "openpyxl" or "csv"
            - row_count: int — total data rows across all sheets
            - page_count: int — alias for sheet_count (compatibility)

    Raises:
        DocumentExtractionError: If the file cannot be read or parsed
    """
    file_path = Path(filepath)
    if not file_path.exists():
        raise DocumentExtractionError(
            f"Spreadsheet file not found: {filepath}",
            details={"filepath": filepath},
        )

    suffix = file_path.suffix.lower()

    if suffix in (".xlsx", ".xls"):
        markdown, sheet_count, row_count, tables_found = _extract_xlsx(file_path)
        method = "openpyxl"
    elif suffix in (".csv", ".tsv"):
        markdown, sheet_count, row_count, tables_found = _extract_csv(file_path)
        method = "csv"
    else:
        raise DocumentExtractionError(
            f"Unsupported spreadsheet format: {suffix}",
            details={"filepath": filepath, "suffix": suffix},
        )

    # Workbook-level cap: multi-sheet files can individually pass the per-sheet
    # cap yet still blow the TELUS gateway budget in aggregate. When the combined
    # markdown exceeds MAX_CHARS_PER_WORKBOOK, keep the prefix that fits and
    # annotate the cut so the LLM (and the reviewer) sees the truncation.
    if len(markdown) > MAX_CHARS_PER_WORKBOOK:
        original_len = len(markdown)
        cut_at = markdown.rfind("\n\n", 0, MAX_CHARS_PER_WORKBOOK)
        if cut_at <= 0:
            cut_at = MAX_CHARS_PER_WORKBOOK
        markdown = markdown[:cut_at] + (
            f"\n\n*Truncated by workbook cap ({MAX_CHARS_PER_WORKBOOK:,} chars): "
            f"kept first {cut_at:,} of {original_len:,} total chars across "
            f"{sheet_count} sheet(s). Full data remains on disk at "
            f"{file_path.name}.*"
        )

    result = {
        "filepath": filepath,
        "markdown": markdown,
        "sheet_count": sheet_count,
        "total_chars": len(markdown),
        "tables_found": tables_found,
        "extracted_at": datetime.now(timezone.utc).isoformat(),
        "extraction_method": method,
        "row_count": row_count,
        "page_count": sheet_count,  # compatibility with PDF-centric code
    }

    logger.info(
        f"Extracted {file_path.name}: {sheet_count} sheet(s), "
        f"{row_count:,} rows, {len(markdown):,} chars"
    )

    return result
