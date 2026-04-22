"""Large-sheet row cap in spreadsheet_extractor (Phase D).

CZ_Plots_LPIS_Evidence_Clean is 128 MB on disk but only 1,254 rows, so
``MAX_ROWS_PER_SHEET`` does not actually trigger on the H1 CSSCP fixture.
The cap still matters for future fixtures — landholdings workbooks with
tens of thousands of plots, remote-sensing index tables, etc. — so it
deserves an explicit regression test.

Two tests:

  * a synthetic workbook with ``MAX_ROWS_PER_SHEET + 1`` data rows must be
    truncated, with a visible footer explaining how many rows were omitted;
  * a workbook exactly at the cap must return every row (no footer).
"""

from __future__ import annotations

import pytest

from registry_review_mcp.extractors.spreadsheet_extractor import (
    MAX_CHARS_PER_SHEET,
    MAX_CHARS_PER_WORKBOOK,
    MAX_ROWS_PER_SHEET,
    extract_spreadsheet,
)


def _write_synth_xlsx(path, row_count: int) -> None:
    """Create an XLSX with one sheet of ``row_count`` data rows plus a header."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plots"
    ws.append(["plot_id", "hectares", "crop"])
    for i in range(row_count):
        ws.append([f"plot-{i:07d}", round(1.0 + (i % 50) * 0.37, 2), "wheat" if i % 2 else "barley"])
    wb.save(str(path))
    wb.close()


@pytest.mark.asyncio
async def test_large_sheet_row_cap_triggers_with_footer(tmp_path):
    """Sheet larger than MAX_ROWS_PER_SHEET must be row-truncated + annotated.

    The char cap may ALSO fire on a long skinny sheet of this size; the
    assertions here cover only the row-cap behavior and tolerate either
    cap order.
    """
    path = tmp_path / "big.xlsx"
    _write_synth_xlsx(path, MAX_ROWS_PER_SHEET + 500)

    result = await extract_spreadsheet(str(path))

    # row_count reflects the post-row-cap count (set before char cap).
    assert result["row_count"] == MAX_ROWS_PER_SHEET
    md = result["markdown"]
    assert "Truncated by row cap: showing first" in md
    assert f"{MAX_ROWS_PER_SHEET:,}" in md
    assert f"{MAX_ROWS_PER_SHEET + 500:,}" in md


@pytest.mark.asyncio
async def test_sheet_below_caps_has_no_truncation(tmp_path):
    """Small sheet must produce no truncation footer at all."""
    path = tmp_path / "small.xlsx"
    # 500 rows × 3 narrow cols is well under both caps.
    _write_synth_xlsx(path, 500)

    result = await extract_spreadsheet(str(path))

    assert result["row_count"] == 500
    md = result["markdown"]
    assert "Truncated" not in md
    assert "plot-0000000" in md
    assert "plot-0000499" in md


def _write_wide_xlsx(path, row_count: int, cols_per_row: int, text_width: int) -> None:
    """Create an XLSX whose rendered markdown will exceed MAX_CHARS_PER_SHEET.

    ``cols_per_row`` and ``text_width`` control how quickly the rendered table
    grows. Row count stays well below MAX_ROWS so the row-cap does NOT trigger;
    the char-cap must be the thing that fires.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "FatRows"
    headers = [f"col_{i}" for i in range(cols_per_row)]
    ws.append(headers)
    filler = "x" * text_width
    for i in range(row_count):
        ws.append([f"r{i}_{filler}" for _ in range(cols_per_row)])
    wb.save(str(path))
    wb.close()


@pytest.mark.asyncio
async def test_char_cap_truncates_oversized_sheet_with_footer(tmp_path):
    """Sheet whose rendered markdown exceeds MAX_CHARS_PER_SHEET must be trimmed.

    Critical guard for the TELUS GPT-OSS gateway: payloads above ~150K chars
    trigger Cloudflare 504/524 retries that poison wall-clock. Cap the sheet,
    annotate the footer, and keep the LLM prompt tractable.
    """
    path = tmp_path / "fat.xlsx"
    # 2500 rows × 20 cols × 200-char filler = ~10M raw chars, well above the cap.
    _write_wide_xlsx(path, row_count=2500, cols_per_row=20, text_width=200)

    result = await extract_spreadsheet(str(path))

    md = result["markdown"]
    assert result["row_count"] == 2500
    # A 10M-char sheet must get trimmed. The workbook cap bounds final size
    # (per-sheet cap fires first but the workbook trim finalises).
    assert len(md) <= MAX_CHARS_PER_WORKBOOK + 2048
    # Either per-sheet or per-workbook truncation footer must be present.
    assert "Truncated by" in md


@pytest.mark.asyncio
async def test_char_cap_does_not_fire_on_small_sheet(tmp_path):
    """Normal-sized sheets must emit full markdown with no char-cap footer."""
    path = tmp_path / "small.xlsx"
    _write_wide_xlsx(path, row_count=10, cols_per_row=4, text_width=20)

    result = await extract_spreadsheet(str(path))

    md = result["markdown"]
    assert "Truncated by char cap" not in md
    assert "Truncated by workbook cap" not in md
    assert len(md) < MAX_CHARS_PER_SHEET


def _write_multi_sheet_xlsx(path, sheet_count: int, rows_per_sheet: int) -> None:
    """Create an XLSX with several sheets each individually under the per-sheet
    cap, but whose combined markdown still exceeds MAX_CHARS_PER_WORKBOOK.

    Used to validate that the workbook-level cap fires when per-sheet caps do not.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    # Remove the default sheet so we can add in a known order
    default = wb.active
    if default is not None:
        wb.remove(default)

    filler = "y" * 40
    for s in range(sheet_count):
        ws = wb.create_sheet(f"sheet_{s}")
        headers = [f"col_{i}" for i in range(6)]
        ws.append(headers)
        for i in range(rows_per_sheet):
            ws.append([f"s{s}r{i}_{filler}" for _ in range(6)])
    wb.save(str(path))
    wb.close()


@pytest.mark.asyncio
async def test_workbook_cap_truncates_multi_sheet_payload(tmp_path):
    """Multi-sheet workbook exceeds MAX_CHARS_PER_WORKBOOK → workbook-level cap fires.

    Each sheet individually sits under the per-sheet cap, but their combined
    markdown would break the TELUS Cloudflare edge. The workbook cap keeps
    the prompt tractable.
    """
    path = tmp_path / "multi.xlsx"
    # 5 sheets × 300 rows each — per sheet well under MAX_CHARS_PER_SHEET (40K),
    # but combined sums past MAX_CHARS_PER_WORKBOOK (60K).
    _write_multi_sheet_xlsx(path, sheet_count=5, rows_per_sheet=300)

    result = await extract_spreadsheet(str(path))

    md = result["markdown"]
    assert result["sheet_count"] == 5
    assert len(md) <= MAX_CHARS_PER_WORKBOOK + 1024
    assert "Truncated by workbook cap" in md
    assert f"{MAX_CHARS_PER_WORKBOOK:,}" in md
