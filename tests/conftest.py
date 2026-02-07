"""Pytest configuration and shared fixtures.

CRITICAL SECURITY: Environment-level test isolation.
XDG environment variables are set BEFORE any package imports to ensure
the settings singleton initializes with /tmp paths from the very start.
This prevents the dual-singleton bug that caused production data loss.
"""

# =============================================================================
# ENVIRONMENT ISOLATION - MUST BE FIRST (before any package imports)
# =============================================================================
import os
import sys
import tempfile

# Check if we're running under pytest
_PYTEST_RUNNING = "pytest" in sys.modules or "py.test" in sys.modules

if _PYTEST_RUNNING:
    # Create isolated test directory BEFORE any settings imports
    _TEST_ROOT = tempfile.mkdtemp(prefix="pytest-registry-")
    os.environ["XDG_DATA_HOME"] = os.path.join(_TEST_ROOT, "data")
    os.environ["XDG_CACHE_HOME"] = os.path.join(_TEST_ROOT, "cache")

    # Ensure directories exist
    os.makedirs(os.environ["XDG_DATA_HOME"], exist_ok=True)
    os.makedirs(os.environ["XDG_CACHE_HOME"], exist_ok=True)

# =============================================================================
# Now safe to import from our package - settings will see /tmp paths
# =============================================================================
import pytest
import pytest_asyncio
import json
from pathlib import Path
from datetime import datetime

# Load cost control plugin
pytest_plugins = ["tests.plugins.cost_control"]

from registry_review_mcp.config.settings import settings, Settings
from registry_review_mcp.extractors.llm_extractors import (
    DateExtractor,
    LandTenureExtractor,
    ProjectIDExtractor,
)
from registry_review_mcp.utils.safe_delete import safe_rmtree


# Global cost tracker for aggregating all test costs
_test_suite_costs = {
    'total_cost': 0.0,
    'total_calls': 0,
    'total_input_tokens': 0,
    'total_output_tokens': 0,
    'cached_calls': 0,
    'by_test': {},
    'start_time': None,
    'end_time': None,
}


def pytest_configure(config):
    """Called before test session starts."""
    _test_suite_costs['start_time'] = datetime.now()

    # Register custom markers
    config.addinivalue_line("markers", "slow: marks tests as slow (real API calls)")
    config.addinivalue_line("markers", "expensive: marks tests with high API costs")

    # SAFETY VALIDATION: Ensure settings point to /tmp
    # This catches any edge case where environment isolation failed
    if not str(settings.sessions_dir).startswith("/tmp"):
        raise RuntimeError(
            f"SAFETY FAILURE: Test isolation failed!\n"
            f"sessions_dir = {settings.sessions_dir}\n"
            f"Expected /tmp prefix. Production data at risk.\n"
            f"XDG_DATA_HOME = {os.environ.get('XDG_DATA_HOME', 'NOT SET')}"
        )

    # Store test root for cleanup
    config._test_root = _TEST_ROOT if _PYTEST_RUNNING else None


def pytest_sessionfinish(session, exitstatus):
    """Called after whole test run finished."""
    _test_suite_costs['end_time'] = datetime.now()

    # Aggregate costs from all /tmp test tracking files
    cost_files = list(Path('/tmp').glob('test_*.json'))

    if not cost_files:
        return

    for file in cost_files:
        try:
            with open(file) as f:
                data = json.load(f)
                _test_suite_costs['total_cost'] += data.get('total_cost_usd', 0.0)
                _test_suite_costs['total_calls'] += data.get('total_api_calls', 0)
                _test_suite_costs['total_input_tokens'] += data.get('total_input_tokens', 0)
                _test_suite_costs['total_output_tokens'] += data.get('total_output_tokens', 0)

                # Count cached calls
                api_calls = data.get('api_calls', [])
                _test_suite_costs['cached_calls'] += sum(1 for call in api_calls if call.get('cached', False))

                # Track by test
                session_id = data.get('session_id', file.stem)
                _test_suite_costs['by_test'][session_id] = {
                    'cost': data.get('total_cost_usd', 0.0),
                    'calls': data.get('total_api_calls', 0),
                }
        except Exception:
            pass

    # Print summary if we have costs
    if _test_suite_costs['total_calls'] > 0:
        print("\n" + "=" * 80)
        print("API COST SUMMARY")
        print("=" * 80)
        print(f"Total API Calls: {_test_suite_costs['total_calls']:,}")
        print(f"  - Real API calls: {_test_suite_costs['total_calls'] - _test_suite_costs['cached_calls']:,}")
        print(f"  - Cached calls: {_test_suite_costs['cached_calls']:,}")
        print(f"Total Tokens: {_test_suite_costs['total_input_tokens'] + _test_suite_costs['total_output_tokens']:,}")
        print(f"  - Input: {_test_suite_costs['total_input_tokens']:,}")
        print(f"  - Output: {_test_suite_costs['total_output_tokens']:,}")
        print(f"Total Cost: ${_test_suite_costs['total_cost']:.4f}")

        if _test_suite_costs['total_calls'] > 0:
            cache_rate = _test_suite_costs['cached_calls'] / _test_suite_costs['total_calls'] * 100
            print(f"Cache Hit Rate: {cache_rate:.1f}%")

        duration = (_test_suite_costs['end_time'] - _test_suite_costs['start_time']).total_seconds()
        print(f"Test Duration: {duration:.1f}s")
        print("=" * 80)

        # Save to file for later analysis
        report_path = Path('test_costs_report.json')
        with open(report_path, 'w') as f:
            json.dump({
                'timestamp': _test_suite_costs['start_time'].isoformat(),
                'duration_seconds': duration,
                'total_cost_usd': _test_suite_costs['total_cost'],
                'total_api_calls': _test_suite_costs['total_calls'],
                'cached_calls': _test_suite_costs['cached_calls'],
                'total_input_tokens': _test_suite_costs['total_input_tokens'],
                'total_output_tokens': _test_suite_costs['total_output_tokens'],
                'by_test': _test_suite_costs['by_test'],
            }, f, indent=2, default=str)

        print(f"\nDetailed cost report saved to: test_costs_report.json")
        print(f"For analysis, run: python scripts/analyze_test_costs.py\n")


# =============================================================================
# Shared Test Fixtures (Cost Optimization)
# =============================================================================


@pytest.fixture(scope="session")
def botany_farm_markdown():
    """Load Botany Farm Project Plan markdown once for all tests.

    Returns the first 20K chars to match what individual tests use.
    """
    project_plan_path = Path("examples/22-23/4997Botany22_Public_Project_Plan/4997Botany22_Public_Project_Plan.md")

    if not project_plan_path.exists():
        pytest.skip("Botany Farm Project Plan not found")

    with open(project_plan_path) as f:
        markdown = f.read()

    # Return first 20K chars (matching what tests use)
    return markdown[:20000]


@pytest_asyncio.fixture(scope="module")
async def botany_farm_dates(botany_farm_markdown):
    """Extract dates from Botany Farm ONCE per test module.

    Module scope is safe because:
    - Data is read-only (tests only read/filter, never mutate)
    - Eliminates expensive duplicate API calls (~$0.03 each)
    - Result is deterministic (same input → same output)
    - Compatible with pytest-asyncio (session scope has issues)

    Requires API key and LLM extraction to be enabled.
    """
    if not settings.anthropic_api_key or not settings.llm_extraction_enabled:
        pytest.skip("LLM extraction not configured")

    extractor = DateExtractor()
    results = await extractor.extract(
        botany_farm_markdown,
        [],
        "Botany_Farm_Project_Plan"
    )
    return results


@pytest_asyncio.fixture(scope="module")
async def botany_farm_tenure(botany_farm_markdown):
    """Extract land tenure from Botany Farm ONCE per test module.

    Module scope is safe because:
    - Data is read-only (tests only read/filter, never mutate)
    - Eliminates expensive duplicate API calls (~$0.03 each)
    - Result is deterministic
    - Compatible with pytest-asyncio (session scope has issues)

    Requires API key and LLM extraction to be enabled.
    """
    if not settings.anthropic_api_key or not settings.llm_extraction_enabled:
        pytest.skip("LLM extraction not configured")

    extractor = LandTenureExtractor()
    results = await extractor.extract(
        botany_farm_markdown,
        [],
        "Botany_Farm_Project_Plan"
    )
    return results


@pytest_asyncio.fixture(scope="module")
async def botany_farm_project_ids(botany_farm_markdown):
    """Extract project IDs from Botany Farm ONCE per test module.

    Module scope is safe because:
    - Data is read-only (tests only read/filter, never mutate)
    - Eliminates expensive duplicate API calls (~$0.03 each)
    - Result is deterministic
    - Compatible with pytest-asyncio (session scope has issues)

    Requires API key and LLM extraction to be enabled.
    """
    if not settings.anthropic_api_key or not settings.llm_extraction_enabled:
        pytest.skip("LLM extraction not configured")

    extractor = ProjectIDExtractor()
    results = await extractor.extract(
        botany_farm_markdown,
        [],
        "Botany_Farm_Project_Plan"
    )
    return results


# =============================================================================
# Original Test Fixtures (Required by existing tests)
# =============================================================================


@pytest.fixture
def temp_data_dir(tmp_path):
    """Create a temporary data directory for tests.

    Args:
        tmp_path: Pytest temporary directory fixture

    Yields:
        Path to temporary data directory
    """
    data_dir = tmp_path / "test_data"
    data_dir.mkdir(parents=True)

    # Create subdirectories
    (data_dir / "checklists").mkdir()
    (data_dir / "sessions").mkdir()
    (data_dir / "cache").mkdir()

    yield data_dir

    # Cleanup using safe_rmtree (will work because tmp_path is in /tmp)
    safe_rmtree(data_dir)


@pytest.fixture
def test_settings(temp_data_dir):
    """Create test settings with temporary directories.

    Args:
        temp_data_dir: Temporary data directory fixture

    Returns:
        Settings instance for testing
    """
    return Settings(
        data_dir=temp_data_dir,
        checklists_dir=temp_data_dir / "checklists",
        sessions_dir=temp_data_dir / "sessions",
        cache_dir=temp_data_dir / "cache",
        enable_caching=True,
    )


@pytest.fixture(scope="session", autouse=True)
def cleanup_cache_once():
    """Clean cache once at the start of the test session.

    This prevents test pollution from previous runs while avoiding
    repeated cleanup operations.
    """
    from registry_review_mcp.utils.cache import Cache

    # Clean the default cache directory
    cache = Cache("test_cleanup")
    cache.clear()

    yield  # Run all tests


@pytest.fixture
def cache(test_settings):
    """Create a cache instance for testing.

    Args:
        test_settings: Test settings fixture

    Returns:
        Cache instance
    """
    from registry_review_mcp.utils.cache import Cache

    return Cache("test", cache_dir=test_settings.cache_dir)


@pytest.fixture(autouse=True, scope="function")
def cleanup_sessions(tmp_path_factory):
    """Ensure each test starts with clean session state.

    STRUCTURAL SAFETY: Environment-level isolation is now the primary defense.
    XDG_DATA_HOME is set to /tmp BEFORE settings is imported, so the settings
    singleton initializes with /tmp paths. This fixture provides additional
    per-test cleanup as a secondary defense.
    """
    # SAFETY: Verify structural isolation is working
    # This should never fail with the new architecture, but we check anyway
    if not str(settings.sessions_dir).startswith("/tmp"):
        raise RuntimeError(
            f"STRUCTURAL ISOLATION FAILURE: settings.sessions_dir = {settings.sessions_dir}\n"
            f"Expected /tmp prefix. The environment isolation is not working."
        )

    def cleanup_test_sessions():
        """Clean up sessions in the settings directory using safe_rmtree."""
        if settings.sessions_dir.exists():
            safe_rmtree(settings.sessions_dir)
            settings.sessions_dir.mkdir(exist_ok=True)

    # Cleanup before test
    cleanup_test_sessions()

    try:
        yield  # Run the test
    finally:
        # Cleanup after test
        cleanup_test_sessions()


@pytest.fixture(scope="session")
def example_documents_path():
    """Get path to example documents (session-scoped for efficiency).

    Session scope is safe because:
    - Returns immutable Path object
    - Path resolution is deterministic
    - Used by 18 tests across 5 files

    Returns:
        Absolute path to examples/22-23 directory
    """
    return (Path(__file__).parent.parent / "examples" / "22-23").absolute()


@pytest.fixture
def sample_xlsx(tmp_path):
    """Create a small test .xlsx with two sheets for spreadsheet ingestion tests.

    Sheet 1 — "Farm Data": 5 data rows of farm records.
    Sheet 2 — "Land Tenure": 3 data rows of tenure records.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Farm Data"
    ws1.append(["Farm ID", "Owner", "Hectares", "Status"])
    for i, (owner, ha, status) in enumerate(
        [("Alice", 45.2, "Active"), ("Bob", 32.1, "Active"), ("Carol", 18.7, "Pending"),
         ("Dan", 55.0, "Active"), ("Eve", 27.3, "Inactive")], start=1
    ):
        ws1.append([f"CE-{i:03d}", owner, ha, status])

    ws2 = wb.create_sheet("Land Tenure")
    ws2.append(["Farm ID", "Deed Number", "Registry", "Date"])
    ws2.append(["CE-001", "D-10234", "County A", "2024-01-15"])
    ws2.append(["CE-002", "D-10235", "County B", "2024-02-20"])
    ws2.append(["CE-003", "D-10236", "County A", "2024-03-10"])

    xlsx_path = tmp_path / "farm_data.xlsx"
    wb.save(str(xlsx_path))
    wb.close()
    return xlsx_path


@pytest.fixture
def sample_csv(tmp_path):
    """Create a small test .csv with monitoring data."""
    import csv as csv_mod

    csv_path = tmp_path / "monitoring_data.csv"
    with open(csv_path, "w", newline="") as f:
        writer = csv_mod.writer(f)
        writer.writerow(["Date", "Plot ID", "SOC (g/kg)", "pH", "Moisture (%)"])
        writer.writerow(["2024-06-01", "P-01", "23.4", "6.8", "31.2"])
        writer.writerow(["2024-06-01", "P-02", "19.7", "7.1", "28.5"])
        writer.writerow(["2024-06-15", "P-01", "24.1", "6.7", "33.0"])
    return csv_path


@pytest.fixture
def sample_land_tenure_xlsx(tmp_path):
    """Create a .xlsx named to trigger land_tenure classification."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenure Records"
    ws.append(["Farm ID", "Deed Number", "Type"])
    ws.append(["CE-001", "D-10234", "Freehold"])

    xlsx_path = tmp_path / "land_tenure_records.xlsx"
    wb.save(str(xlsx_path))
    wb.close()
    return xlsx_path
